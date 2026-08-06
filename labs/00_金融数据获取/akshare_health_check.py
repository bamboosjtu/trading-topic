"""AKShare 股票接口健康检测脚本。

职责仅限于：
1. 从 AKShare 官方股票文档和当前安装包发现股票接口；
2. 逐个执行最小化调用，检查接口在当前环境下是否可运行；
3. 按“数据源 / 接口名称”排序；
4. 输出不可变的 JSON/XLSX 运行快照，并更新 latest 副本。

本脚本不负责业务分类，不会修改：
- akshare_stock_api.xlsx
- akshare_api.md
- akshare_tutorial.ipynb

默认运行方式：
    uv run --project labs python labs/00_金融数据获取/akshare_health_check.py

快速验证前 10 个接口：
    uv run --project labs python labs/00_金融数据获取/akshare_health_check.py --limit 10

仅检测名称匹配的接口：
    uv run --project labs python labs/00_金融数据获取/akshare_health_check.py --match "stock_zh_a"
"""

from __future__ import annotations

import argparse
import hashlib
import inspect
import json
import multiprocessing as mp
import os
import platform
import re
import shutil
import sys
import time
import traceback
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.parse import urlparse

import akshare as ak
import pandas as pd
import requests

try:
    from bs4 import BeautifulSoup
except ImportError:  # pragma: no cover - AKShare 环境通常已间接安装
    BeautifulSoup = None  # type: ignore[assignment]


# -----------------------------------------------------------------------------
# 配置
# -----------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DOC_URL = "https://akshare.akfamily.xyz/data/stock/stock.html"
DEFAULT_OUTPUT_DIR = SCRIPT_DIR / "health"

SNAPSHOT_SCHEMA_VERSION = "1.0"
TEST_CONFIG_VERSION = "1.0"
DEFAULT_TIMEOUT_SECONDS = 15.0

TEST_SYMBOL_A = "000001"       # 平安银行
TEST_SYMBOL_A_SH = "600036"    # 招商银行
TEST_SYMBOL_A_TX = "sz000001"
TEST_SYMBOL_HK = "00700"
TEST_SYMBOL_US = "105.MSFT"

STATUS_AVAILABLE = "AVAILABLE"
STATUS_EMPTY = "EMPTY"
STATUS_FAILED = "FAILED"
STATUS_TIMEOUT = "TIMEOUT"
STATUS_MISSING = "MISSING"
STATUS_SKIPPED = "SKIPPED"

STATUS_ORDER = {
    STATUS_AVAILABLE: 0,
    STATUS_EMPTY: 1,
    STATUS_SKIPPED: 2,
    STATUS_TIMEOUT: 3,
    STATUS_FAILED: 4,
    STATUS_MISSING: 5,
}

SENSITIVE_KEY_RE = re.compile(
    r"token|api[_-]?key|authorization|cookie|password|secret",
    re.IGNORECASE,
)

PROXY_ENV_KEYS = (
    "HTTP_PROXY",
    "HTTPS_PROXY",
    "ALL_PROXY",
    "http_proxy",
    "https_proxy",
    "all_proxy",
    "NO_PROXY",
    "no_proxy",
)

# 文档目标地址优先于函数名后缀，用于得到单一且稳定的数据源字段。
DOMAIN_SOURCE_MAP: tuple[tuple[str, str], ...] = (
    ("eastmoney.com", "东方财富"),
    ("sina.com.cn", "新浪财经"),
    ("sina.com", "新浪财经"),
    ("qq.com", "腾讯财经"),
    ("xueqiu.com", "雪球"),
    ("sse.com.cn", "上海证券交易所"),
    ("szse.cn", "深圳证券交易所"),
    ("bse.cn", "北京证券交易所"),
    ("hkex.com.hk", "香港交易所"),
    ("10jqka.com.cn", "同花顺"),
    ("iwencai.com", "i问财"),
    ("cninfo.com.cn", "巨潮资讯"),
    ("baidu.com", "百度股市通"),
    ("cls.cn", "财联社"),
    ("futunn.com", "富途牛牛"),
    ("futu5.com", "富途牛牛"),
    ("legulegu.com", "乐咕乐股"),
    ("eniu.com", "亿牛网"),
    ("csindex.com.cn", "中证指数"),
    ("swindex.com", "申万指数"),
    ("csrc.gov.cn", "中国证监会"),
    ("stcn.com", "证券时报"),
    ("chinamoney.com.cn", "中国货币网"),
    ("nasdaq.com", "Nasdaq"),
)

# 只覆盖那些通用参数推断容易明显出错的接口。
FUNCTION_ARG_OVERRIDES: dict[str, dict[str, Any]] = {
    "stock_zh_a_hist": {
        "symbol": TEST_SYMBOL_A,
        "period": "daily",
        "adjust": "",
    },
    "stock_zh_a_hist_tx": {
        "symbol": TEST_SYMBOL_A_TX,
        "adjust": "",
    },
    "stock_zh_a_daily": {
        "symbol": TEST_SYMBOL_A_TX,
        "adjust": "",
    },
    "stock_zh_a_hist_min_em": {
        "symbol": TEST_SYMBOL_A,
        "period": "5",
        "adjust": "",
    },
    "stock_zh_a_hist_pre_min_em": {
        "symbol": TEST_SYMBOL_A,
    },
    "stock_hk_hist": {
        "symbol": TEST_SYMBOL_HK,
        "period": "daily",
        "adjust": "",
    },
    "stock_hk_hist_min_em": {
        "symbol": TEST_SYMBOL_HK,
        "period": "5",
        "adjust": "",
    },
    "stock_us_hist": {
        "symbol": TEST_SYMBOL_US,
        "period": "daily",
        "adjust": "",
    },
    "stock_us_hist_min_em": {
        "symbol": TEST_SYMBOL_US,
        "period": "5",
        "adjust": "",
    },
    "stock_sector_spot": {
        "indicator": "行业",
    },
    "stock_sector_detail": {
        "sector": "hangye_ZB07",
    },
}


# -----------------------------------------------------------------------------
# 数据模型
# -----------------------------------------------------------------------------


@dataclass(slots=True)
class InterfaceDefinition:
    """接口目录信息，不包含业务分类。"""

    interface_name: str
    description: str = ""
    target_url: str = ""
    input_params: list[str] = field(default_factory=list)
    upstream: str = "未识别"
    documented: bool = False
    installed: bool = False
    callable: bool = False


@dataclass(slots=True)
class HealthResult:
    """单个接口的一次运行结果。"""

    interface_name: str
    description: str
    upstream: str
    target_url: str
    input_params: list[str]
    documented: bool
    installed: bool
    callable: bool
    status: str
    rows: int = 0
    columns: list[str] = field(default_factory=list)
    result_type: str = ""
    elapsed_seconds: float = 0.0
    error_type: str = ""
    error_message: str = ""
    test_case: str = ""
    test_params_hash: str = ""
    test_args_preview: dict[str, Any] = field(default_factory=dict)
    tested_at: str = ""


# -----------------------------------------------------------------------------
# 终端显示
# -----------------------------------------------------------------------------


class ConsoleStyle:
    RESET = "\033[0m"
    BOLD = "\033[1m"
    DIM = "\033[2m"
    GREEN = "\033[32m"
    YELLOW = "\033[33m"
    RED = "\033[31m"
    MAGENTA = "\033[35m"
    CYAN = "\033[36m"
    GRAY = "\033[90m"

    def __init__(self) -> None:
        self.enabled = sys.stdout.isatty() and "NO_COLOR" not in os.environ

    def apply(self, text: str, *styles: str) -> str:
        if not self.enabled:
            return text
        return "".join(styles) + text + self.RESET


CONSOLE = ConsoleStyle()

STATUS_STYLES = {
    STATUS_AVAILABLE: ConsoleStyle.GREEN,
    STATUS_EMPTY: ConsoleStyle.YELLOW,
    STATUS_FAILED: ConsoleStyle.RED,
    STATUS_TIMEOUT: ConsoleStyle.MAGENTA,
    STATUS_MISSING: ConsoleStyle.GRAY,
    STATUS_SKIPPED: ConsoleStyle.CYAN,
}


def _status_text(status: str) -> str:
    return CONSOLE.apply(status, STATUS_STYLES.get(status, ""), ConsoleStyle.BOLD)


# -----------------------------------------------------------------------------
# 通用工具
# -----------------------------------------------------------------------------


def _now() -> datetime:
    return datetime.now().astimezone()


def _iso_now() -> str:
    return _now().isoformat(timespec="seconds")


def _safe_version(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z._-]+", "_", value)


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _json_default(value: Any) -> Any:
    if isinstance(value, (date, datetime, pd.Timestamp)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (set, tuple)):
        return list(value)
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    return repr(value)


def _json_dumps(value: Any, *, indent: int | None = None) -> str:
    return json.dumps(
        value,
        ensure_ascii=False,
        sort_keys=True,
        indent=indent,
        default=_json_default,
    )


def _sanitize_error(text: str, limit: int = 500) -> str:
    """清理错误信息中的密钥、本地路径和多余空白。"""
    text = text.replace("\r", " ").replace("\n", " | ")
    text = re.sub(
        r"(?i)(token|api[_-]?key|authorization|cookie|password|secret)"
        r"\s*[:=]\s*([^&\s|,;]+)",
        r"\1=***",
        text,
    )
    text = re.sub(r"[A-Za-z]:[\\/][^\s|\"']+", "<local-path>", text)
    text = re.sub(r"(?<![A-Za-z])/(?:Users|home)/[^\s|\"']+", "<local-path>", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[-limit:]


def _safe_args_preview(args: dict[str, Any]) -> dict[str, Any]:
    preview: dict[str, Any] = {}
    for key, value in args.items():
        preview[key] = "***" if SENSITIVE_KEY_RE.search(key) else value
    return preview


def _args_hash(interface_name: str, args: dict[str, Any]) -> str:
    payload = {
        "interface_name": interface_name,
        "args": args,
        "test_config_version": TEST_CONFIG_VERSION,
    }
    return _sha256_text(_json_dumps(payload))[:16]


def _previous_weekday(value: date) -> date:
    while value.weekday() >= 5:
        value -= timedelta(days=1)
    return value


def _test_dates() -> dict[str, str]:
    end = _previous_weekday(date.today() - timedelta(days=1))
    start = end - timedelta(days=35)
    return {
        "start_date": start.strftime("%Y%m%d"),
        "end_date": end.strftime("%Y%m%d"),
        "date": end.strftime("%Y%m%d"),
        "begin_date": start.strftime("%Y%m%d"),
        "year": str(max(end.year - 1, 2000)),
        "quarter": f"{max(end.year - 1, 2000)}1",
        "start_time": start.strftime("%Y%m%d") + "093000",
        "end_time": end.strftime("%Y%m%d") + "150000",
    }


# -----------------------------------------------------------------------------
# 文档解析与接口发现
# -----------------------------------------------------------------------------


def fetch_documentation(url: str, timeout: float = 30.0) -> tuple[str, str]:
    """获取官方股票接口文档；失败时返回空文本和错误信息。"""
    try:
        response = requests.get(url, timeout=timeout)
        response.raise_for_status()
        return response.text, ""
    except Exception as exc:
        return "", _sanitize_error(f"{type(exc).__name__}: {exc}")


def _parse_params_table(table: Any) -> list[str]:
    params: list[str] = []
    rows = table.find_all("tr")
    for row in rows[1:]:
        cells = row.find_all(["td", "th"])
        if not cells:
            continue
        name = cells[0].get_text(strip=True)
        if name and name != "-":
            params.append(name)
    return params


def parse_documentation(html: str) -> list[InterfaceDefinition]:
    """从官方股票文档提取接口名、描述、目标地址和参数。"""
    if not html:
        return []

    if BeautifulSoup is None:
        names = sorted(set(re.findall(r"接口[:：]\s*(stock_[A-Za-z0-9_]+)", html)))
        return [InterfaceDefinition(interface_name=name, documented=True) for name in names]

    soup = BeautifulSoup(html, "html.parser")
    interfaces: list[InterfaceDefinition] = []
    current: InterfaceDefinition | None = None
    expect_params_table = False

    elements = [
        element
        for element in soup.find_all(["p", "table"], recursive=True)
        if not element.find_parent("table")
    ]

    for element in elements:
        tag = element.name
        text = element.get_text(strip=True)

        if tag == "p":
            match = re.match(r"接口[:：]\s*(\S+)", text)
            if match:
                if current is not None:
                    interfaces.append(current)
                current = InterfaceDefinition(
                    interface_name=match.group(1).strip(),
                    documented=True,
                )
                expect_params_table = False
                continue

            if current is None:
                continue

            match = re.match(r"目标地址[:：]\s*(\S+)", text)
            if match:
                current.target_url = match.group(1).strip()
                continue

            match = re.match(r"描述[:：]\s*(.+)", text)
            if match:
                current.description = match.group(1).strip()
                continue

            if text == "输入参数":
                expect_params_table = True
                continue

        if tag == "table" and expect_params_table and current is not None:
            current.input_params = _parse_params_table(element)
            expect_params_table = False

    if current is not None:
        interfaces.append(current)

    # 文档偶尔重复展示同一接口，按接口名保留信息更完整的一条。
    deduplicated: dict[str, InterfaceDefinition] = {}
    for item in interfaces:
        previous = deduplicated.get(item.interface_name)
        if previous is None:
            deduplicated[item.interface_name] = item
            continue
        if len(item.description) > len(previous.description):
            previous.description = item.description
        if item.target_url and not previous.target_url:
            previous.target_url = item.target_url
        if len(item.input_params) > len(previous.input_params):
            previous.input_params = item.input_params

    return list(deduplicated.values())


def _source_from_domain(target_url: str) -> str | None:
    if not target_url:
        return None
    try:
        parsed = urlparse(target_url if "://" in target_url else f"https://{target_url}")
        host = (parsed.hostname or "").lower()
    except Exception:
        host = target_url.lower()

    for domain, source in DOMAIN_SOURCE_MAP:
        if domain in host:
            return source
    return None


def _source_from_name(interface_name: str, description: str) -> str | None:
    name = interface_name.lower()
    tokens = set(name.split("_"))
    text = f"{interface_name} {description}".lower()

    suffix_rules: tuple[tuple[Callable[[], bool], str], ...] = (
        (lambda: "em" in tokens or name.endswith("_em"), "东方财富"),
        (lambda: "sina" in tokens or "新浪" in description, "新浪财经"),
        (lambda: "tx" in tokens or "腾讯" in description, "腾讯财经"),
        (lambda: "xq" in tokens or "雪球" in description, "雪球"),
        (lambda: "ths" in tokens or "同花顺" in description, "同花顺"),
        (lambda: "iwencai" in tokens or "问财" in description, "i问财"),
        (lambda: "cninfo" in tokens or "巨潮" in description, "巨潮资讯"),
        (lambda: "baidu" in tokens or "百度" in description, "百度股市通"),
        (lambda: "futu" in tokens or "富途" in description, "富途牛牛"),
        (lambda: "eniu" in tokens or "亿牛" in description, "亿牛网"),
        (lambda: "lg" in tokens or "乐咕" in description, "乐咕乐股"),
        (lambda: "sse" in tokens or "上交所" in description, "上海证券交易所"),
        (lambda: "szse" in tokens or "深交所" in description, "深圳证券交易所"),
        (lambda: "bse" in tokens or "北交所" in description, "北京证券交易所"),
        (lambda: "cls" in tokens or "财联社" in description, "财联社"),
    )

    for predicate, source in suffix_rules:
        if predicate():
            return source

    if "通达信" in text or "mootdx" in text:
        return "通达信"
    return None


def guess_upstream(interface_name: str, target_url: str, description: str) -> str:
    return (
        _source_from_domain(target_url)
        or _source_from_name(interface_name, description)
        or "未识别"
    )


def discover_interfaces(
    documented: Iterable[InterfaceDefinition],
) -> list[InterfaceDefinition]:
    """合并官方文档接口与当前 AKShare 安装包中的公开 stock_* 函数。"""
    catalog: dict[str, InterfaceDefinition] = {
        item.interface_name: item for item in documented
    }

    installed_stock_names = {
        name
        for name in dir(ak)
        if name.startswith("stock_")
        and not name.startswith("stock__")
        and callable(getattr(ak, name, None))
    }

    for name in installed_stock_names:
        if name not in catalog:
            catalog[name] = InterfaceDefinition(interface_name=name)

    for name, item in catalog.items():
        attribute = getattr(ak, name, None)
        item.installed = attribute is not None
        item.callable = callable(attribute)
        item.upstream = guess_upstream(name, item.target_url, item.description)

    return sorted(
        catalog.values(),
        key=lambda item: (item.upstream, item.interface_name),
    )


# -----------------------------------------------------------------------------
# 测试参数构造
# -----------------------------------------------------------------------------


def _has_default(parameter: inspect.Parameter) -> bool:
    return parameter.default is not inspect.Parameter.empty


def _required_parameters(signature: inspect.Signature) -> list[inspect.Parameter]:
    return [
        parameter
        for parameter in signature.parameters.values()
        if parameter.kind
        not in (inspect.Parameter.VAR_POSITIONAL, inspect.Parameter.VAR_KEYWORD)
        and not _has_default(parameter)
    ]


def _symbol_for_interface(interface_name: str) -> str:
    name = interface_name.lower()
    if "_hk_" in name or name.startswith("stock_hk"):
        return TEST_SYMBOL_HK
    if "_us_" in name or name.startswith("stock_us"):
        return TEST_SYMBOL_US
    if name in {"stock_zh_a_daily", "stock_zh_a_hist_tx"}:
        return TEST_SYMBOL_A_TX
    if any(token in name for token in ("minute", "intraday", "tick", "bid_ask")):
        return TEST_SYMBOL_A_SH
    return TEST_SYMBOL_A


def _value_for_required_parameter(
    interface_name: str,
    parameter: inspect.Parameter,
    dates: dict[str, str],
) -> tuple[bool, Any]:
    name = parameter.name
    lower_interface = interface_name.lower()

    if name in {"symbol", "stock", "code", "security"}:
        return True, _symbol_for_interface(interface_name)
    if name in dates:
        return True, dates[name]
    if name == "start_year":
        return True, dates["year"]
    if name == "end_year":
        return True, str(int(dates["year"]) + 1)
    if name == "period":
        if "minute" in lower_interface or "intraday" in lower_interface:
            return True, "5"
        return True, "daily"
    if name == "adjust":
        return True, ""
    if name == "market":
        return True, "沪"
    if name == "exchange":
        return True, "上海证券交易所"
    if name == "indicator":
        return False, None
    if name in {"page", "page_no", "page_num", "from_page", "to_page"}:
        return True, 1
    if name in {"page_size", "limit", "size", "count"}:
        return True, 20
    if name == "timeout":
        return True, 10
    if name == "quarter":
        return True, dates["quarter"]
    if name == "year":
        return True, dates["year"]
    if name == "sector":
        return True, "hangye_ZB07"

    return False, None


def build_test_args(interface_name: str, function: Any) -> tuple[dict[str, Any] | None, str]:
    """构造最小化测试参数；无法可靠推断时返回 SKIPPED 原因。"""
    try:
        signature = inspect.signature(function)
    except (TypeError, ValueError) as exc:
        return None, f"无法读取函数签名: {type(exc).__name__}: {exc}"

    parameters = signature.parameters
    dates = _test_dates()
    args: dict[str, Any] = {}

    # 明确覆盖先应用，并只保留当前函数实际接受的参数。
    for key, value in FUNCTION_ARG_OVERRIDES.get(interface_name, {}).items():
        if key in parameters:
            args[key] = value

    # 用较短的时间范围和较小分页降低接口压力，但不覆盖有明确默认值的业务参数。
    safe_optional_overrides: dict[str, Any] = {
        "start_date": dates["start_date"],
        "end_date": dates["end_date"],
        "begin_date": dates["begin_date"],
        "date": dates["date"],
        "start_time": dates["start_time"],
        "end_time": dates["end_time"],
        "from_page": 1,
        "to_page": 1,
        "page": 1,
        "page_no": 1,
        "page_num": 1,
        "page_size": 20,
        "limit": 20,
        "timeout": 10,
    }
    for key, value in safe_optional_overrides.items():
        if key in parameters and key not in args:
            args[key] = value

    unresolved: list[str] = []
    for parameter in _required_parameters(signature):
        if parameter.name in args:
            continue
        resolved, value = _value_for_required_parameter(interface_name, parameter, dates)
        if resolved:
            args[parameter.name] = value
        else:
            unresolved.append(parameter.name)

    if unresolved:
        return None, "缺少可靠测试值的必填参数: " + ", ".join(unresolved)

    # 不向只接受 *args/**kwargs 的包装函数注入猜测参数。
    concrete_parameters = [
        parameter
        for parameter in parameters.values()
        if parameter.kind
        not in (inspect.Parameter.VAR_POSITIONAL, inspect.Parameter.VAR_KEYWORD)
    ]
    if not concrete_parameters:
        return {}, "default"

    return args, "default"


# -----------------------------------------------------------------------------
# 子进程调用与健康判定
# -----------------------------------------------------------------------------


def _disable_proxy_for_akshare() -> None:
    """仅在 AKShare 测试子进程中绕过系统代理。"""
    for key in PROXY_ENV_KEYS:
        os.environ.pop(key, None)
    os.environ["NO_PROXY"] = "*"
    os.environ["no_proxy"] = "*"


def _result_metadata(result: Any) -> dict[str, Any]:
    result_type = type(result).__name__
    rows = 0
    columns: list[str] = []

    if isinstance(result, pd.DataFrame):
        rows = len(result)
        columns = [str(column) for column in result.columns[:50]]
    elif isinstance(result, pd.Series):
        rows = len(result)
        if result.name is not None:
            columns = [str(result.name)]
        else:
            columns = [str(index) for index in result.index[:50]]
    elif isinstance(result, dict):
        rows = len(result)
        columns = [str(key) for key in list(result.keys())[:50]]
    elif isinstance(result, (list, tuple, set)):
        rows = len(result)
        first = next(iter(result), None)
        if isinstance(first, dict):
            columns = [str(key) for key in list(first.keys())[:50]]
    elif result is None:
        rows = 0
    else:
        rows = 1

    return {
        "ok": True,
        "rows": rows,
        "columns": columns,
        "result_type": result_type,
        "error_type": "",
        "error_message": "",
    }


def _process_worker(
    send_connection: Any,
    interface_name: str,
    args: dict[str, Any],
) -> None:
    """子进程入口：执行一个 AKShare 接口并只回传轻量元数据。"""
    try:
        _disable_proxy_for_akshare()
        function = getattr(ak, interface_name)
        result = function(**args)
        payload = _result_metadata(result)
    except BaseException as exc:  # 子进程必须将接口异常转换为结构化结果
        payload = {
            "ok": False,
            "rows": 0,
            "columns": [],
            "result_type": "",
            "error_type": type(exc).__name__,
            "error_message": _sanitize_error(
                "".join(traceback.format_exception_only(type(exc), exc))
            ),
        }

    try:
        send_connection.send(payload)
    finally:
        send_connection.close()


def _direct_call(interface_name: str, args: dict[str, Any]) -> dict[str, Any]:
    """调试模式：在当前进程直接调用，不提供硬超时。"""
    try:
        original_env = {key: os.environ.get(key) for key in PROXY_ENV_KEYS}
        _disable_proxy_for_akshare()
        try:
            result = getattr(ak, interface_name)(**args)
        finally:
            for key in PROXY_ENV_KEYS:
                os.environ.pop(key, None)
            for key, value in original_env.items():
                if value is not None:
                    os.environ[key] = value
        return _result_metadata(result)
    except BaseException as exc:
        return {
            "ok": False,
            "rows": 0,
            "columns": [],
            "result_type": "",
            "error_type": type(exc).__name__,
            "error_message": _sanitize_error(
                "".join(traceback.format_exception_only(type(exc), exc))
            ),
        }


def execute_with_timeout(
    interface_name: str,
    args: dict[str, Any],
    timeout_seconds: float,
    mode: str,
) -> tuple[dict[str, Any] | None, float, bool]:
    """返回 (子进程结果, 耗时, 是否超时)。"""
    started = time.perf_counter()

    if mode == "direct":
        payload = _direct_call(interface_name, args)
        return payload, time.perf_counter() - started, False

    context = mp.get_context("spawn")
    receive_connection, send_connection = context.Pipe(duplex=False)
    process = context.Process(
        target=_process_worker,
        args=(send_connection, interface_name, args),
        daemon=True,
    )

    process.start()
    send_connection.close()
    process.join(timeout_seconds)

    if process.is_alive():
        process.terminate()
        process.join(3.0)
        receive_connection.close()
        return None, time.perf_counter() - started, True

    payload: dict[str, Any] | None = None
    if receive_connection.poll(1.0):
        try:
            payload = receive_connection.recv()
        except EOFError:
            payload = None
    receive_connection.close()

    if payload is None:
        payload = {
            "ok": False,
            "rows": 0,
            "columns": [],
            "result_type": "",
            "error_type": "WorkerExitError",
            "error_message": f"测试子进程异常退出，exitcode={process.exitcode}",
        }

    return payload, time.perf_counter() - started, False


def test_interface(
    definition: InterfaceDefinition,
    timeout_seconds: float,
    mode: str,
) -> HealthResult:
    tested_at = _iso_now()

    if not definition.installed or not definition.callable:
        return HealthResult(
            **asdict(definition),
            status=STATUS_MISSING,
            error_type="MissingInterface",
            error_message="接口在当前 AKShare 安装包中不存在或不可调用",
            tested_at=tested_at,
        )

    function = getattr(ak, definition.interface_name)
    args, test_case = build_test_args(definition.interface_name, function)

    if args is None:
        return HealthResult(
            **asdict(definition),
            status=STATUS_SKIPPED,
            error_type="MissingTestArguments",
            error_message=_sanitize_error(test_case),
            test_case="signature_unresolved",
            tested_at=tested_at,
        )

    test_params_hash = _args_hash(definition.interface_name, args)
    preview = _safe_args_preview(args)
    payload, elapsed, timed_out = execute_with_timeout(
        definition.interface_name,
        args,
        timeout_seconds,
        mode,
    )

    if timed_out:
        return HealthResult(
            **asdict(definition),
            status=STATUS_TIMEOUT,
            elapsed_seconds=round(elapsed, 3),
            error_type="TimeoutError",
            error_message=f"超过 {timeout_seconds:.1f} 秒硬超时",
            test_case=test_case,
            test_params_hash=test_params_hash,
            test_args_preview=preview,
            tested_at=tested_at,
        )

    assert payload is not None
    if not payload.get("ok", False):
        return HealthResult(
            **asdict(definition),
            status=STATUS_FAILED,
            elapsed_seconds=round(elapsed, 3),
            error_type=str(payload.get("error_type", "Exception")),
            error_message=_sanitize_error(str(payload.get("error_message", ""))),
            test_case=test_case,
            test_params_hash=test_params_hash,
            test_args_preview=preview,
            tested_at=tested_at,
        )

    rows = int(payload.get("rows", 0) or 0)
    status = STATUS_AVAILABLE if rows > 0 else STATUS_EMPTY
    return HealthResult(
        **asdict(definition),
        status=status,
        rows=rows,
        columns=[str(column) for column in payload.get("columns", [])],
        result_type=str(payload.get("result_type", "")),
        elapsed_seconds=round(elapsed, 3),
        test_case=test_case,
        test_params_hash=test_params_hash,
        test_args_preview=preview,
        tested_at=tested_at,
    )


def run_health_checks(
    definitions: list[InterfaceDefinition],
    timeout_seconds: float,
    mode: str,
) -> list[HealthResult]:
    total = len(definitions)
    results: list[HealthResult] = []
    current_source: str | None = None

    print()
    print(CONSOLE.apply("AKShare 股票接口健康检查", ConsoleStyle.BOLD))
    print(f"AKShare: {ak.__version__}  接口数: {total}  模式: {mode}")
    print(f"单接口超时: {timeout_seconds:.1f}s  测试时间: {_iso_now()}")

    for index, definition in enumerate(definitions, start=1):
        if definition.upstream != current_source:
            current_source = definition.upstream
            print()
            print(CONSOLE.apply(f"[{current_source}]", ConsoleStyle.CYAN, ConsoleStyle.BOLD))

        result = test_interface(definition, timeout_seconds, mode)
        results.append(result)
        row_text = f"{result.rows} 行" if result.rows else "-"
        print(
            f"  {index:>3}/{total:<3} "
            f"{result.interface_name:<42} "
            f"{_status_text(result.status):<18} "
            f"{row_text:>8}  {result.elapsed_seconds:>7.3f}s"
        )

    return sorted(results, key=lambda item: (item.upstream, item.interface_name))


# -----------------------------------------------------------------------------
# 快照输出
# -----------------------------------------------------------------------------


def _status_counts(results: list[HealthResult]) -> dict[str, int]:
    counts = {status: 0 for status in STATUS_ORDER}
    for result in results:
        counts[result.status] = counts.get(result.status, 0) + 1
    return counts


def _source_counts(results: list[HealthResult]) -> list[dict[str, Any]]:
    frame = pd.DataFrame(
        {
            "数据源": [result.upstream for result in results],
            "状态": [result.status for result in results],
        }
    )
    if frame.empty:
        return []

    pivot = pd.crosstab(frame["数据源"], frame["状态"])
    for status in STATUS_ORDER:
        if status not in pivot.columns:
            pivot[status] = 0
    pivot = pivot[list(STATUS_ORDER.keys())]
    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_index()
    return pivot.reset_index().to_dict(orient="records")


def build_snapshot(
    results: list[HealthResult],
    *,
    run_id: str,
    started_at: str,
    finished_at: str,
    duration_seconds: float,
    documentation_url: str,
    documentation_hash: str,
    documentation_error: str,
    timeout_seconds: float,
    mode: str,
) -> dict[str, Any]:
    status_counts = _status_counts(results)
    config = {
        "test_config_version": TEST_CONFIG_VERSION,
        "timeout_seconds": timeout_seconds,
        "execution_mode": mode,
        "proxy_policy": "AKShare 子进程内 NO_PROXY=*；文档请求沿用系统环境",
        "test_dates": _test_dates(),
        "function_arg_overrides": FUNCTION_ARG_OVERRIDES,
    }
    config_hash = _sha256_text(_json_dumps(config))

    return {
        "snapshot_schema_version": SNAPSHOT_SCHEMA_VERSION,
        "run_id": run_id,
        "metadata": {
            "akshare_version": str(ak.__version__),
            "python_version": platform.python_version(),
            "platform": platform.platform(),
            "started_at": started_at,
            "finished_at": finished_at,
            "duration_seconds": round(duration_seconds, 3),
            "documentation_url": documentation_url,
            "documentation_hash": documentation_hash,
            "documentation_error": documentation_error,
            "test_config_version": TEST_CONFIG_VERSION,
            "test_config_hash": config_hash,
            "execution_mode": mode,
            "timeout_seconds": timeout_seconds,
        },
        "summary": {
            "interface_count": len(results),
            "documented_count": sum(result.documented for result in results),
            "installed_count": sum(result.installed for result in results),
            "callable_count": sum(result.callable for result in results),
            "status_counts": status_counts,
            "source_counts": _source_counts(results),
        },
        "interfaces": [asdict(result) for result in results],
    }


def write_snapshot_json(snapshot: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(_json_dumps(snapshot, indent=2) + "\n", encoding="utf-8")
    temporary.replace(path)


def _interface_rows(snapshot: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in snapshot["interfaces"]:
        rows.append(
            {
                "数据源": item["upstream"],
                "接口名称": item["interface_name"],
                "描述": item["description"],
                "状态": item["status"],
                "文档存在": "是" if item["documented"] else "否",
                "安装包存在": "是" if item["installed"] else "否",
                "可调用": "是" if item["callable"] else "否",
                "返回行数": item["rows"],
                "返回类型": item["result_type"],
                "返回字段": ", ".join(item["columns"]),
                "耗时(秒)": item["elapsed_seconds"],
                "错误类型": item["error_type"],
                "错误信息": item["error_message"],
                "测试参数摘要": _json_dumps(item["test_args_preview"]),
                "测试参数哈希": item["test_params_hash"],
                "目标地址": item["target_url"],
                "文档输入参数": ", ".join(item["input_params"]),
                "测试时间": item["tested_at"],
            }
        )
    return rows


def _apply_excel_style(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="D9EAF7")
    status_fills = {
        STATUS_AVAILABLE: PatternFill("solid", fgColor="C6EFCE"),
        STATUS_EMPTY: PatternFill("solid", fgColor="FFEB9C"),
        STATUS_FAILED: PatternFill("solid", fgColor="FFC7CE"),
        STATUS_TIMEOUT: PatternFill("solid", fgColor="E4DFEC"),
        STATUS_MISSING: PatternFill("solid", fgColor="D9D9D9"),
        STATUS_SKIPPED: PatternFill("solid", fgColor="DDEBF7"),
    }

    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_index, cells in enumerate(worksheet.columns, start=1):
            values = [str(cell.value) if cell.value is not None else "" for cell in cells]
            width = min(max(max((len(value) for value in values), default=0) + 2, 10), 50)
            worksheet.column_dimensions[get_column_letter(column_index)].width = width

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary_sheet = workbook["运行摘要"]
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 70
    for cell in summary_sheet["A"]:
        if cell.row > 1:
            cell.fill = title_fill
            cell.font = Font(bold=True)

    detail_sheet = workbook["接口明细"]
    headers = {cell.value: cell.column for cell in detail_sheet[1]}
    status_column = headers.get("状态")
    interface_column = headers.get("接口名称")
    error_column = headers.get("错误信息")

    if interface_column:
        for row in range(2, detail_sheet.max_row + 1):
            detail_sheet.cell(row=row, column=interface_column).font = Font(name="Consolas")

    if error_column:
        detail_sheet.column_dimensions[get_column_letter(error_column)].width = 60

    if status_column:
        status_letter = get_column_letter(status_column)
        for status, fill in status_fills.items():
            formula = f'${status_letter}2="{status}"'
            detail_sheet.conditional_formatting.add(
                f"A2:R{detail_sheet.max_row}",
                FormulaRule(formula=[formula], fill=fill),
            )

    workbook.save(path)


def write_snapshot_xlsx(snapshot: dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    metadata = snapshot["metadata"]
    summary = snapshot["summary"]
    status_counts = summary["status_counts"]

    summary_rows = [
        {"项目": "运行 ID", "值": snapshot["run_id"]},
        {"项目": "AKShare 版本", "值": metadata["akshare_version"]},
        {"项目": "Python 版本", "值": metadata["python_version"]},
        {"项目": "操作系统", "值": metadata["platform"]},
        {"项目": "开始时间", "值": metadata["started_at"]},
        {"项目": "完成时间", "值": metadata["finished_at"]},
        {"项目": "总耗时(秒)", "值": metadata["duration_seconds"]},
        {"项目": "接口总数", "值": summary["interface_count"]},
        {"项目": "文档接口数", "值": summary["documented_count"]},
        {"项目": "安装包接口数", "值": summary["installed_count"]},
        {"项目": "可调用接口数", "值": summary["callable_count"]},
        *(
            {"项目": status, "值": count}
            for status, count in status_counts.items()
        ),
        {"项目": "单接口超时(秒)", "值": metadata["timeout_seconds"]},
        {"项目": "执行模式", "值": metadata["execution_mode"]},
        {"项目": "测试配置版本", "值": metadata["test_config_version"]},
        {"项目": "测试配置哈希", "值": metadata["test_config_hash"]},
        {"项目": "官方文档", "值": metadata["documentation_url"]},
        {"项目": "文档哈希", "值": metadata["documentation_hash"]},
        {"项目": "文档获取错误", "值": metadata["documentation_error"]},
    ]

    interface_frame = pd.DataFrame(_interface_rows(snapshot))
    interface_frame = interface_frame.sort_values(
        ["数据源", "接口名称"],
        kind="stable",
        na_position="last",
    )
    source_frame = pd.DataFrame(summary["source_counts"])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="运行摘要", index=False)
        source_frame.to_excel(writer, sheet_name="数据源统计", index=False)
        interface_frame.to_excel(writer, sheet_name="接口明细", index=False)

    _apply_excel_style(path)


def write_outputs(snapshot: dict[str, Any], output_dir: Path) -> tuple[Path, Path]:
    run_id = snapshot["run_id"]
    version = _safe_version(snapshot["metadata"]["akshare_version"])
    runs_dir = output_dir / "runs"
    runs_dir.mkdir(parents=True, exist_ok=True)

    stem = f"{run_id}_akshare_{version}"
    json_path = runs_dir / f"{stem}.json"
    xlsx_path = runs_dir / f"{stem}.xlsx"

    write_snapshot_json(snapshot, json_path)
    write_snapshot_xlsx(snapshot, xlsx_path)

    output_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(json_path, output_dir / "latest.json")
    shutil.copy2(xlsx_path, output_dir / "latest.xlsx")

    return json_path, xlsx_path


def print_summary(snapshot: dict[str, Any], json_path: Path, xlsx_path: Path) -> None:
    counts = snapshot["summary"]["status_counts"]
    print()
    print(CONSOLE.apply("检测完成", ConsoleStyle.BOLD))
    print("  " + "  ".join(f"{status}={counts.get(status, 0)}" for status in STATUS_ORDER))
    print(f"  JSON: {json_path}")
    print(f"  XLSX: {xlsx_path}")
    print(f"  latest: {json_path.parent.parent / 'latest.json'}")


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="检测当前环境中 AKShare 股票接口的运行健康状态。",
    )
    parser.add_argument(
        "--doc-url",
        default=DOC_URL,
        help="AKShare 股票接口官方文档地址。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="健康快照输出目录。",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=DEFAULT_TIMEOUT_SECONDS,
        help="单个接口硬超时秒数，默认 15。",
    )
    parser.add_argument(
        "--mode",
        choices=("process", "direct"),
        default="process",
        help="process 提供硬超时；direct 较快但无法终止卡住的请求。",
    )
    parser.add_argument(
        "--match",
        default="",
        help="只检测接口名匹配该正则表达式的接口。",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="仅检测排序后的前 N 个接口；0 表示不限制。",
    )
    parser.add_argument(
        "--skip-doc",
        action="store_true",
        help="不请求官方文档，只检测当前安装包公开的 stock_* 接口。",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.timeout <= 0:
        print("--timeout 必须大于 0", file=sys.stderr)
        return 2
    if args.limit < 0:
        print("--limit 不能小于 0", file=sys.stderr)
        return 2

    started = _now()
    started_at = started.isoformat(timespec="seconds")
    run_id = started.strftime("%Y%m%d_%H%M%S")

    print(CONSOLE.apply("=" * 72, ConsoleStyle.DIM))
    print(CONSOLE.apply("AKShare 股票接口健康检测", ConsoleStyle.BOLD))
    print(CONSOLE.apply("=" * 72, ConsoleStyle.DIM))

    html = ""
    documentation_error = ""
    if not args.skip_doc:
        print(f"读取官方文档: {args.doc_url}")
        html, documentation_error = fetch_documentation(args.doc_url)
        if documentation_error:
            print(CONSOLE.apply(f"文档读取失败，将仅使用安装包接口: {documentation_error}", ConsoleStyle.YELLOW))
        else:
            print(f"文档大小: {len(html):,} 字符")

    documented = parse_documentation(html)
    definitions = discover_interfaces(documented)

    if args.match:
        try:
            pattern = re.compile(args.match)
        except re.error as exc:
            print(f"--match 正则无效: {exc}", file=sys.stderr)
            return 2
        definitions = [item for item in definitions if pattern.search(item.interface_name)]

    if args.limit:
        definitions = definitions[: args.limit]

    if not definitions:
        print("未发现可检测的股票接口。", file=sys.stderr)
        return 1

    print(
        f"发现接口: {len(definitions)} 个 "
        f"(文档 {sum(item.documented for item in definitions)}, "
        f"安装包 {sum(item.installed for item in definitions)})"
    )

    results = run_health_checks(
        definitions,
        timeout_seconds=args.timeout,
        mode=args.mode,
    )

    finished = _now()
    snapshot = build_snapshot(
        results,
        run_id=run_id,
        started_at=started_at,
        finished_at=finished.isoformat(timespec="seconds"),
        duration_seconds=(finished - started).total_seconds(),
        documentation_url=args.doc_url,
        documentation_hash=_sha256_text(html) if html else "",
        documentation_error=documentation_error,
        timeout_seconds=args.timeout,
        mode=args.mode,
    )

    json_path, xlsx_path = write_outputs(snapshot, args.output_dir.resolve())
    print_summary(snapshot, json_path, xlsx_path)
    return 0


if __name__ == "__main__":
    mp.freeze_support()
    raise SystemExit(main())
