"""
对 docs/tutorial/akshare_api.xlsx 中的 375 个接口按用途进行分类。

分类原则:
1. 按投资研究实际工作流设计大类(从市场总览 → 标的筛选 → 个股深研 → 资金筹码 → 事件公告 → 衍生主题)。
2. 以"接口函数名"和真实用途为归类依据,不机械沿用原表"接口名称"首段(原表存在归类错位,例如
   序号 113-118 标在"沪深港通持股"下但实际是停复牌/分红/新闻/报告时间;序号 318-320 标在
   "北京证券交易所-标的证券信息"下但实际是盈利预测;序号 246-249 标在"机构推荐"下但实际是
   担保/诉讼/质押/目标价)。
3. 每个接口按函数名精确匹配归入唯一大类,边界规则在 RULES 中显式定义。
4. 同时输出 Markdown 报告(便于阅读)和 Excel(便于检索)。
"""
import json
import re
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"C:/Users/theTruth/Documents/projects/vibe-working/trading-topic")
DUMP = ROOT / ".workbuddy" / "akshare_api_dump.json"
MD_OUT = ROOT / "docs" / "tutorial" / "akshare_api_classification.md"
XLSX_OUT = ROOT / "docs" / "tutorial" / "akshare_api_classified.xlsx"

data = json.loads(DUMP.read_text(encoding="utf-8"))


# ============================================================
# 分类规则:按函数名 (func) 精确匹配,顺序敏感,先命中先归。
# 每条规则:(大类key, 子类key, 关键词列表-任一命中即归此)
# ============================================================
# 大类定义(顺序即展示顺序)
CATEGORIES = [
    ("market_overview",  "一、市场总览与统计"),
    ("listing_info",     "二、标的列表与基础信息"),
    ("quote",            "三、行情数据"),
    ("finance_value",    "四、财务与估值"),
    ("shareholder",      "五、股东与股本变动"),
    ("capital_flow",     "六、资金与筹码"),
    ("ipo_capital",      "七、IPO 与资本运作"),
    ("institution",      "八、机构与研究"),
    ("announcement",     "九、公告与事件异动"),
    ("sentiment",        "十、市场情绪、互动与 ESG"),
]

# 子类定义
SUBCATS = {
    "market_overview": [
        ("市场总貌与每日概况", ["stock_sse_summary", "stock_szse_summary", "stock_szse_area_summary",
                              "stock_szse_sector_summary", "stock_sse_deal_daily"]),
        ("账户与活跃度统计", ["stock_account_statistics_em", "stock_market_activity_legu"]),
        ("大盘估值水位", ["stock_a_ttm_lyr", "stock_a_all_pb", "stock_market_pe_lg",
                       "stock_market_pb_lg", "stock_index_pe_lg", "stock_index_pb_lg",
                       "stock_a_high_low_statistics", "stock_a_below_net_asset_statistics",
                       "stock_a_gxl_lg", "stock_hk_gxl_lg", "stock_a_congestion_lg",
                       "stock_ebs_lg", "stock_buffett_index_lg"]),
    ],
    "listing_info": [
        ("股票列表与代码字典", ["stock_info_a_code_name", "stock_info_sh_name_code",
                             "stock_info_sz_name_code", "stock_info_bj_name_code",
                             "stock_info_sz_delist", "stock_staq_net_stop",
                             "stock_info_sh_delist", "stock_zh_ah_name"]),
        ("个股基础信息", ["stock_individual_info_em", "stock_individual_basic_info_xq",
                       "stock_individual_basic_info_us_xq", "stock_individual_basic_info_hk_xq",
                       "stock_hk_security_profile_em", "stock_hk_company_profile_em",
                       "stock_profile_cninfo", "stock_ipo_summary_cninfo"]),
        ("股票更名与名称变更", ["stock_info_change_name", "stock_info_sz_change_name"]),
        ("行业/板块分类", ["stock_industry_category_cninfo", "stock_industry_change_cninfo",
                       "stock_industry_clf_hist_sw", "stock_industry_pe_ratio_cninfo",
                       "stock_board_concept_name_em", "stock_board_concept_info_ths",
                       "stock_board_industry_name_em", "stock_board_industry_summary_ths"]),
    ],
    "quote": [
        ("A股实时行情", ["stock_zh_a_spot_em", "stock_sh_a_spot_em", "stock_sz_a_spot_em",
                      "stock_bj_a_spot_em", "stock_new_a_spot_em", "stock_cy_a_spot_em",
                      "stock_kc_a_spot_em", "stock_zh_ab_comparison_em", "stock_zh_a_spot",
                      "stock_individual_spot_xq", "stock_bid_ask_em", "stock_zh_kcb_spot",
                      "stock_zh_a_st_em", "stock_zh_a_new_em", "stock_zh_a_stop_em",
                      "stock_zh_a_new", "stock_zh_a_cdr_daily"]),
        ("B股/科创板/创业板实时行情", ["stock_zh_b_spot_em", "stock_zh_b_spot",
                                   "stock_zh_b_daily", "stock_zh_b_minute",
                                   "stock_zh_kcb_daily", "stock_zh_kcb_report_em"]),
        ("A股历史行情与分时", ["stock_zh_a_hist", "stock_zh_a_daily", "stock_zh_a_hist_tx",
                           "stock_zh_a_minute", "stock_zh_a_hist_min_em", "stock_intraday_em",
                           "stock_intraday_sina", "stock_zh_a_hist_pre_min_em",
                           "stock_zh_a_tick_tx"]),
        ("AH股行情", ["stock_zh_ah_spot_em", "stock_zh_ah_spot", "stock_zh_ah_daily"]),
        ("港股行情", ["stock_hk_spot_em", "stock_hk_main_board_spot_em", "stock_hk_spot",
                   "stock_hk_hist_min_em", "stock_hk_hist", "stock_hk_daily",
                   "stock_hk_famous_spot_em"]),
        ("美股行情", ["stock_us_spot_em", "stock_us_spot", "stock_us_hist",
                   "stock_us_hist_min_em", "stock_us_daily", "stock_us_pink_spot_em",
                   "stock_us_famous_spot_em"]),
        ("概念板块行情", ["stock_board_concept_spot_em", "stock_board_concept_cons_em",
                      "stock_board_concept_hist_em", "stock_board_concept_hist_min_em",
                      "stock_concept_cons_futu", "stock_board_concept_index_ths"]),
        ("行业板块行情", ["stock_board_industry_spot_em", "stock_board_industry_cons_em",
                      "stock_board_industry_hist_em", "stock_board_industry_hist_min_em",
                      "stock_board_industry_index_ths"]),
        ("新浪板块行情", ["stock_sector_spot", "stock_sector_detail"]),
    ],
    "finance_value": [
        ("三大报表-东财", ["stock_balance_sheet_by_report_em", "stock_balance_sheet_by_yearly_em",
                      "stock_profit_sheet_by_report_em", "stock_profit_sheet_by_yearly_em",
                      "stock_profit_sheet_by_quarterly_em", "stock_cash_flow_sheet_by_report_em",
                      "stock_cash_flow_sheet_by_yearly_em", "stock_cash_flow_sheet_by_quarterly_em",
                      "stock_balance_sheet_by_report_delisted_em",
                      "stock_profit_sheet_by_report_delisted_em",
                      "stock_cash_flow_sheet_by_report_delisted_em",
                      "stock_zcfz_em", "stock_zcfz_bj_em", "stock_lrb_em", "stock_xjll_em"]),
        ("三大报表-同花顺/新浪/港股/美股", ["stock_financial_debt_new_ths", "stock_financial_benefit_new_ths",
                                  "stock_financial_cash_new_ths", "stock_financial_report_sina",
                                  "stock_financial_hk_report_em", "stock_financial_us_report_em"]),
        ("财务指标与关键指标", ["stock_financial_abstract", "stock_financial_abstract_new_ths",
                          "stock_financial_analysis_indicator_em", "stock_financial_analysis_indicator",
                          "stock_financial_hk_analysis_indicator_em",
                          "stock_financial_us_analysis_indicator_em",
                          "stock_hk_financial_indicator_em", "stock_hk_indicator_eniu"]),
        ("业绩报表/快报/预告/披露时间", ["stock_yjbb_em", "stock_yjkb_em", "stock_yjyg_em",
                                "stock_yysj_em", "stock_report_disclosure",
                                "stock_zh_a_disclosure_report_cninfo",
                                "stock_zh_a_disclosure_relation_cninfo"]),
        ("估值指标", ["stock_zh_valuation_baidu", "stock_value_em",
                  "stock_hk_valuation_baidu", "stock_us_valuation_baidu"]),
        ("同行/行业对比", ["stock_zh_growth_comparison_em", "stock_zh_valuation_comparison_em",
                      "stock_zh_dupont_comparison_em", "stock_zh_scale_comparison_em",
                      "stock_hk_growth_comparison_em", "stock_hk_valuation_comparison_em",
                      "stock_hk_scale_comparison_em"]),
        ("商誉", ["stock_sy_profile_em", "stock_sy_yq_em", "stock_sy_jz_em",
               "stock_sy_em", "stock_sy_hy_em"]),
        ("港股分红派息(财务口径)", ["stock_hk_dividend_payout_em"]),
    ],
    "shareholder": [
        ("十大股东/十大流通股东", ["stock_gdfx_free_top_10_em", "stock_gdfx_top_10_em",
                          "stock_gdfx_free_holding_change_em", "stock_gdfx_holding_change_em",
                          "stock_gdfx_free_holding_analyse_em", "stock_gdfx_holding_analyse_em",
                          "stock_gdfx_free_holding_detail_em", "stock_gdfx_holding_detail_em",
                          "stock_gdfx_free_holding_statistics_em", "stock_gdfx_holding_statistics_em",
                          "stock_gdfx_free_holding_teamwork_em", "stock_gdfx_holding_teamwork_em",
                          "stock_shareholder_change_ths"]),
        ("股东户数", ["stock_zh_a_gdhs", "stock_zh_a_gdhs_detail_em", "stock_hold_num_cninfo"]),
        ("高管与董监高持股变动", ["stock_management_change_ths", "stock_share_hold_change_sse",
                          "stock_share_hold_change_szse", "stock_share_hold_change_bse",
                          "stock_hold_management_detail_cninfo", "stock_hold_management_detail_em",
                          "stock_hold_management_person_em"]),
        ("股本结构与股本变动", ["stock_zh_a_gbjg_em", "stock_share_change_cninfo",
                       "stock_hold_change_cninfo", "stock_allotment_cninfo"]),
        ("主要股东/流通股东/实控人", ["stock_main_stock_holder", "stock_circulate_stock_holder",
                          "stock_hold_control_cninfo", "stock_ggcg_em"]),
    ],
    "capital_flow": [
        ("个股/大盘资金流", ["stock_fund_flow_individual", "stock_fund_flow_concept",
                      "stock_fund_flow_industry", "stock_fund_flow_big_deal",
                      "stock_individual_fund_flow", "stock_individual_fund_flow_rank",
                      "stock_market_fund_flow", "stock_main_fund_flow"]),
        ("板块资金流", ["stock_sector_fund_flow_rank", "stock_sector_fund_flow_summary",
                  "stock_sector_fund_flow_hist", "stock_concept_fund_flow_hist"]),
        ("沪深港通-持股与统计", ["stock_hsgt_individual_em", "stock_hsgt_individual_detail_em",
                         "stock_hsgt_hold_stock_em", "stock_hsgt_stock_statistics_em",
                         "stock_hsgt_institution_statistics_em", "stock_hsgt_sh_hk_spot_em",
                         "stock_hsgt_hist_em", "stock_hsgt_board_rank_em",
                         "stock_hsgt_fund_flow_summary_em", "stock_hsgt_fund_min_em",
                         "stock_hk_ggt_components_em"]),
        ("沪深港通-汇率", ["stock_sgt_settlement_exchange_rate_szse",
                     "stock_sgt_settlement_exchange_rate_sse",
                     "stock_sgt_reference_exchange_rate_szse",
                     "stock_sgt_reference_exchange_rate_sse"]),
        ("融资融券", ["stock_margin_ratio_pa", "stock_margin_account_info",
                 "stock_margin_sse", "stock_margin_detail_sse", "stock_margin_szse",
                 "stock_margin_detail_szse", "stock_margin_underlying_info_szse",
                 "stock_margin_bse", "stock_margin_detail_bse",
                 "stock_margin_underlying_info_bse", "stock_yzxdr_em"]),
        ("龙虎榜", ["stock_lhb_detail_em", "stock_lhb_stock_statistic_em",
                "stock_lhb_jgmmtj_em", "stock_lhb_jgstatistic_em", "stock_lhb_hyyyb_em",
                "stock_lhb_yyb_detail_em", "stock_lhb_yybph_em", "stock_lhb_traderstatistic_em",
                "stock_lhb_stock_detail_em", "stock_lh_yyb_most", "stock_lh_yyb_capital",
                "stock_lh_yyb_control", "stock_lhb_detail_daily_sina", "stock_lhb_ggtj_sina",
                "stock_lhb_yytj_sina", "stock_lhb_jgzz_sina", "stock_lhb_jgmx_sina"]),
        ("大宗交易", ["stock_dzjy_sctj", "stock_dzjy_mrmx", "stock_dzjy_mrtj",
                 "stock_dzjy_hygtj", "stock_dzjy_hyyybtj", "stock_dzjy_yybph"]),
        ("股权质押", ["stock_gpzy_profile_em", "stock_gpzy_pledge_ratio_em",
                 "stock_gpzy_pledge_ratio_detail_em", "stock_gpzy_individual_pledge_ratio_detail_em",
                 "stock_gpzy_distribute_statistics_company_em",
                 "stock_gpzy_distribute_statistics_bank_em", "stock_gpzy_industry_data_em",
                 "stock_cg_equity_mortgage_cninfo"]),
        ("主力控盘与市场热度(筹码视角)", ["stock_comment_detail_zlkp_jgcyd_em",
                               "stock_comment_detail_zhpj_lspf_em",
                               "stock_comment_detail_scrd_focus_em",
                               "stock_comment_detail_scrd_desire_em",
                               "stock_cyq_em"]),
    ],
    "ipo_capital": [
        ("新股发行与申购", ["stock_ipo_info", "stock_dxsyl_em", "stock_xgsglb_em",
                      "stock_ipo_ths", "stock_ipo_hk_ths", "stock_new_ipo_cninfo",
                      "stock_new_gh_cninfo", "stock_xgsr_ths", "stock_ipo_benefit_ths"]),
        ("IPO审核与辅导", ["stock_ipo_review_em", "stock_ipo_tutor_em", "stock_ipo_declare_em",
                      "stock_register_all_em", "stock_register_kcb", "stock_register_cyb",
                      "stock_register_sh", "stock_register_sz", "stock_register_bj",
                      "stock_register_db"]),
        ("增发/配股/回购", ["stock_add_stock", "stock_qbzf_em", "stock_pg_em",
                     "stock_repurchase_em"]),
        ("限售解禁", ["stock_restricted_release_queue_sina", "stock_restricted_release_summary_em",
                 "stock_restricted_release_detail_em", "stock_restricted_release_queue_em",
                 "stock_restricted_release_stockholder_em"]),
        ("分红与历史分红", ["stock_fhps_em", "stock_fhps_detail_em", "stock_fhps_detail_ths",
                    "stock_hk_fhpx_detail_ths", "stock_history_dividend",
                    "stock_history_dividend_detail", "stock_dividend_cninfo"]),
        ("股东大会", ["stock_gddh_em"]),
    ],
    "institution": [
        ("机构调研", ["stock_jgdy_tj_em", "stock_jgdy_detail_em", "stock_zyjs_ths", "stock_zygc_em"]),
        ("机构持股", ["stock_fund_stock_holder", "stock_institute_hold",
                 "stock_institute_hold_detail", "stock_report_fund_hold",
                 "stock_report_fund_hold_detail"]),
        ("机构推荐与评级", ["stock_institute_recommend", "stock_institute_recommend_detail",
                    "stock_rank_forecast_cninfo"]),
        ("个股研报", ["stock_research_report_em"]),
        ("分析师", ["stock_analyst_rank_em", "stock_analyst_detail_em", "stock_comment_em"]),
        ("盈利预测", ["stock_profit_forecast_em", "stock_hk_profit_forecast_et",
                 "stock_profit_forecast_ths"]),
        ("券商业绩", ["stock_qsjy_em"]),
    ],
    "announcement": [
        ("公告与信息披露", ["stock_notice_report", "stock_individual_notice_report"]),
        ("重大合同", ["stock_zdhtmx_em"]),
        ("公司动态", ["stock_gsrl_gsdt_em"]),
        ("停复牌与除权除息通知", ["stock_tfp_em", "news_trade_notify_suspend_baidu",
                          "news_trade_notify_dividend_baidu"]),
        ("异动股池", ["stock_zt_pool_em", "stock_zt_pool_previous_em", "stock_zt_pool_strong_em",
                 "stock_zt_pool_sub_new_em", "stock_zt_pool_zbgc_em",
                 "stock_zt_pool_dtgc_em", "stock_changes_em"]),
        ("板块/成份变动", ["stock_board_change_em"]),
        ("内部交易与险资举牌", ["stock_inner_trade_xq", "stock_rank_xzjp_ths"]),
        ("对外担保/公司诉讼", ["stock_cg_guarantee_cninfo", "stock_cg_lawsuit_cninfo"]),
        ("技术形态选股", ["stock_rank_cxfl_ths", "stock_rank_cxsl_ths", "stock_rank_xstp_ths",
                    "stock_rank_xxtp_ths", "stock_rank_ljqs_ths", "stock_rank_ljqd_ths"]),
        ("报告披露时间", ["news_report_time_baidu"]),
    ],
    "sentiment": [
        ("股票热度(雪球/东财)", ["stock_hot_follow_xq", "stock_hot_tweet_xq", "stock_hot_deal_xq",
                          "stock_hot_rank_em", "stock_hot_up_em", "stock_hk_hot_rank_em",
                          "stock_hot_rank_detail_em", "stock_hk_hot_rank_detail_em",
                          "stock_hot_rank_detail_realtime_em",
                          "stock_hk_hot_rank_detail_realtime_em",
                          "stock_hot_rank_latest_em", "stock_hk_hot_rank_latest_em",
                          "stock_hot_keyword_em", "stock_hot_search_baidu",
                          "stock_hot_rank_relate_em"]),
        ("涨跌投票", ["stock_zh_vote_baidu"]),
        ("互动平台", ["stock_irm_cninfo", "stock_irm_ans_cninfo", "stock_sns_sseinfo"]),
        ("新闻", ["stock_news_em", "stock_news_main_cx"]),
        ("美港目标价", ["stock_price_js"]),
        ("ESG 评级", ["stock_esg_rate_sina", "stock_esg_msci_sina", "stock_esg_rft_sina",
                 "stock_esg_zd_sina", "stock_esg_hz_sina"]),
    ],
}

# ============================================================
# 归类执行
# ============================================================
# 建立 func -> 接口对象 的映射
by_func = {d["接口"]: d for d in data}

classified = {ck: defaultdict(list) for ck, _ in CATEGORIES}
used_funcs = set()

for cat_key, _ in CATEGORIES:
    for sub_name, funcs in SUBCATS[cat_key]:
        for f in funcs:
            if f in by_func and f not in used_funcs:
                classified[cat_key][sub_name].append(by_func[f])
                used_funcs.add(f)

# 未归类
unclassified = [d for d in data if d["接口"] not in used_funcs]

# 校验
total_classified = sum(len(lst) for cat in classified.values() for lst in cat.values())
print(f"已分类: {total_classified} / 总数: {len(data)} / 未归类: {len(unclassified)}")
if unclassified:
    print("\n=== 未归类接口 ===")
    for d in unclassified:
        print(f"  #{d['序号']:>3d} [{d['可用性']}] {d['接口名称']}  ||  {d['接口']}")

# ============================================================
# 输出 Markdown
# ============================================================
def avail_badge(a: str) -> str:
    return "✅可用" if a == "可用" else "❌不可用"

md = []
md.append("# AkShare 股票接口分类总览\n")
md.append(f"> 数据来源:`docs/tutorial/akshare_api.xlsx`(工作表「AkShare 股票接口测试」)\n")
md.append(f"> 接口总数:**{len(data)}** 个  |  可用:**{sum(1 for d in data if d['可用性']=='可用')}** 个  |  不可用:**{sum(1 for d in data if d['可用性']!='可用')}** 个\n")
md.append(f"> 分类依据:以**接口函数名与真实用途**为准(原表「接口名称」首段存在归类错位,已校正)\n")
md.append("")
md.append("## 分类速览\n")
md.append("| 大类 | 子类数 | 接口数 | 可用 | 不可用 |")
md.append("| --- | ---: | ---: | ---: | ---: |")
for cat_key, cat_name in CATEGORIES:
    sub_count = len(classified[cat_key])
    items = [d for sub in classified[cat_key].values() for d in sub]
    ok = sum(1 for d in items if d["可用性"] == "可用")
    no = len(items) - ok
    md.append(f"| {cat_name} | {sub_count} | {len(items)} | {ok} | {no} |")
if unclassified:
    md.append(f"| (未归类) | - | {len(unclassified)} | {sum(1 for d in unclassified if d['可用性']=='可用')} | {sum(1 for d in unclassified if d['可用性']!='可用')} |")
md.append("")

for cat_key, cat_name in CATEGORIES:
    md.append(f"## {cat_name}\n")
    for sub_name, items in classified[cat_key].items():
        md.append(f"### {sub_name}  ({len(items)})\n")
        md.append("| 序号 | 接口名称 | 接口函数 | 数据源 | 输入参数 | 可用性 |")
        md.append("| ---: | --- | --- | --- | --- | :---: |")
        for d in sorted(items, key=lambda x: x["序号"]):
            name = str(d["接口名称"]).replace("|", "\\|")
            func = str(d["接口"]).replace("|", "\\|")
            src = str(d["数据源"]).replace("|", "\\|")
            params = str(d["输入参数"]).replace("|", "\\|") if d["输入参数"] else "-"
            md.append(f"| {d['序号']} | {name} | `{func}` | {src} | {params} | {avail_badge(d['可用性'])} |")
        md.append("")
    md.append("")

if unclassified:
    md.append("## 附:未归类接口\n")
    md.append("| 序号 | 接口名称 | 接口函数 | 数据源 | 可用性 |")
    md.append("| ---: | --- | --- | --- | :---: |")
    for d in unclassified:
        md.append(f"| {d['序号']} | {d['接口名称']} | `{d['接口']}` | {d['数据源']} | {avail_badge(d['可用性'])} |")
    md.append("")

md.append("## 使用说明\n")
md.append("- **可用性** 列标注 ✅ 的接口可直接调用;标注 ❌ 的接口在原表中标为「不可用」,使用前需先到 AkShare 文档核对是否已修复或改名。\n")
md.append("- **数据源** 主要集中在东方财富(207)、新浪财经(43)、同花顺(32)、巨潮资讯(23)、沪深北交易所等,跨源同主题接口可互为备份。\n")
md.append("- 同一主题(如三大报表、龙虎榜、沪深港通持股)在多个数据源下均有覆盖,做研究时建议优先选**可用**且**字段最全**的版本,再做交叉校验。\n")
md.append("- 本分类按投资研究工作流组织:从**市场总览**到**标的筛选**(列表/基础信息),再到**个股深研**(行情/财务/股东),接着**资金筹码**(资金流/陆股通/两融/龙虎榜/质押),然后**IPO 与资本运作**、**机构研究**、**公告事件**,最后**情绪/互动/ESG** 等衍生主题。\n")
md.append("- 如需按数据源或可用性筛选,可打开配套文件 `docs/tutorial/akshare_api_classified.xlsx`,在工作表中按列筛选。\n")

MD_OUT.write_text("\n".join(md), encoding="utf-8")
print(f"\nMarkdown 已写入: {MD_OUT}")

# ============================================================
# 输出 Excel
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = "接口分类"

header_fill = PatternFill("solid", fgColor="2F5496")
header_font = Font(bold=True, color="FFFFFF", size=11)
cat_fill = PatternFill("solid", fgColor="D9E2F3")
sub_fill = PatternFill("solid", fgColor="F2F2F2")
ok_fill = PatternFill("solid", fgColor="E2EFDA")
no_fill = PatternFill("solid", fgColor="FCE4D6")
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

headers = ["大类", "子类", "序号", "接口名称", "接口函数", "数据源", "目标地址", "输入参数", "可用性"]
ws.append(headers)
for col_idx, _ in enumerate(headers, 1):
    c = ws.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border

for cat_key, cat_name in CATEGORIES:
    for sub_name, items in classified[cat_key].items():
        for d in sorted(items, key=lambda x: x["序号"]):
            row = [
                cat_name, sub_name, d["序号"], str(d["接口名称"]), str(d["接口"]),
                str(d["数据源"]), str(d["目标地址"]) if d["目标地址"] else "",
                str(d["输入参数"]) if d["输入参数"] else "", str(d["可用性"]),
            ]
            ws.append(row)
            r = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col_idx)
                cell.border = border
                cell.alignment = left if col_idx in (2, 4, 5, 7, 8) else center
            # 可用性着色
            avail_cell = ws.cell(row=r, column=9)
            avail_cell.fill = ok_fill if d["可用性"] == "可用" else no_fill

if unclassified:
    for d in unclassified:
        row = ["(未归类)", "-", d["序号"], str(d["接口名称"]), str(d["接口"]),
               str(d["数据源"]), str(d["目标地址"]) if d["目标地址"] else "",
               str(d["输入参数"]) if d["输入参数"] else "", str(d["可用性"])]
        ws.append(row)
        r = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.border = border
            cell.alignment = left if col_idx in (2, 4, 5, 7, 8) else center
        avail_cell = ws.cell(row=r, column=9)
        avail_cell.fill = ok_fill if d["可用性"] == "可用" else no_fill

# 列宽
widths = [22, 26, 8, 38, 42, 14, 42, 18, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

# 添加一个"分类统计"工作表
ws2 = wb.create_sheet("分类统计")
ws2.append(["大类", "子类数", "接口数", "可用", "不可用"])
for col_idx in range(1, 6):
    c = ws2.cell(row=1, column=col_idx)
    c.fill = header_fill
    c.font = header_font
    c.alignment = center
    c.border = border
for cat_key, cat_name in CATEGORIES:
    sub_count = len(classified[cat_key])
    items = [d for sub in classified[cat_key].values() for d in sub]
    ok = sum(1 for d in items if d["可用性"] == "可用")
    no = len(items) - ok
    ws2.append([cat_name, sub_count, len(items), ok, no])
    r = ws2.max_row
    for col_idx in range(1, 6):
        cell = ws2.cell(row=r, column=col_idx)
        cell.border = border
        cell.alignment = left if col_idx == 1 else center
if unclassified:
    ok = sum(1 for d in unclassified if d["可用性"] == "可用")
    ws2.append(["(未归类)", 0, len(unclassified), ok, len(unclassified) - ok])
ws2.append(["合计", "-", len(data), sum(1 for d in data if d["可用性"] == "可用"),
            sum(1 for d in data if d["可用性"] != "可用")])
r = ws2.max_row
for col_idx in range(1, 6):
    cell = ws2.cell(row=r, column=col_idx)
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = left if col_idx == 1 else center
for i, w in enumerate([22, 10, 10, 10, 10], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

wb.save(XLSX_OUT)
print(f"Excel 已写入: {XLSX_OUT}")

# 控制台汇总
print("\n=== 分类汇总 ===")
for cat_key, cat_name in CATEGORIES:
    sub_count = len(classified[cat_key])
    items = [d for sub in classified[cat_key].values() for d in sub]
    ok = sum(1 for d in items if d["可用性"] == "可用")
    print(f"{cat_name}: 子类 {sub_count} / 接口 {len(items)} (可用 {ok})")
if unclassified:
    print(f"(未归类): {len(unclassified)}")
