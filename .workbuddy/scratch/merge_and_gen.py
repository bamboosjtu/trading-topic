"""Merge cleaned interface names with test results, then regenerate Excel."""
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

INTERFACES_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\.workbuddy\scratch\interfaces.json"
RESULTS_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\.workbuddy\scratch\results.json"
OUT_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\akshare.api.xlsx"

with open(INTERFACES_PATH, "r", encoding="utf-8") as f:
    interfaces = json.load(f)
with open(RESULTS_PATH, "r", encoding="utf-8") as f:
    results = json.load(f)

# Build a lookup of test status by func name
status_by_func = {r["func"]: r.get("status", "不可用") for r in results}

# Use the cleaned interface data as the source of truth for names/sources/params,
# and pull the test status from results.
merged = []
for iface in interfaces:
    merged.append({
        "name": iface["name"],
        "data_source": iface["data_source"],
        "func": iface["func"],
        "url": iface["url"],
        "params_raw": iface["params_raw"] or "null",
        "status": status_by_func.get(iface["func"], "不可用"),
    })

# Generate Excel
wb = Workbook()
ws = wb.active
ws.title = "AkShare 股票接口测试"

headers = ["序号", "接口名称", "数据源", "接口", "目标地址", "输入参数", "可用性"]
ws.append(headers)

header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ok_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
cell_font = Font(name="微软雅黑", size=10)
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, r in enumerate(merged, 1):
    params = r["params_raw"] if r["params_raw"] and r["params_raw"].lower() != "null" else "null"
    row = [
        i,
        r["name"],
        r["data_source"],
        r["func"],
        r["url"],
        params,
        r["status"],
    ]
    ws.append(row)
    row_idx = i + 1
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font
        cell.border = border
        if col_idx in (1, 7):
            cell.alignment = center_align
        else:
            cell.alignment = cell_align
    status_cell = ws.cell(row=row_idx, column=7)
    if status_cell.value == "可用":
        status_cell.fill = ok_fill
    else:
        status_cell.fill = fail_fill

widths = [6, 34, 16, 38, 52, 28, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 24

wb.save(OUT_PATH)

total = len(merged)
ok = sum(1 for r in merged if r["status"] == "可用")
print(f"Saved: {OUT_PATH}")
print(f"Total: {total}, 可用: {ok}, 不可用: {total - ok}")
