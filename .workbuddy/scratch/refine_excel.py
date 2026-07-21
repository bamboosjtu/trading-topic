"""Refine akshare.api.xlsx:
1. Delete unavailable interfaces (keep only 可用)
2. Add a 类别 column (after 序号, before 接口名称) based on the 10 categories
   defined in docs/tutorial/akshare_api_classification.md
3. Renumber 序号
"""
import json
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CLASS_DOC = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\docs\tutorial\akshare_api_classification.md"
RESULTS_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\.workbuddy\scratch\results.json"
INTERFACES_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\.workbuddy\scratch\interfaces.json"
OUT_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\akshare.api.xlsx"

# ---- Step 1: Parse the classification doc to build func -> category mapping ----
with open(CLASS_DOC, "r", encoding="utf-8") as f:
    md = f.read()

func_to_category = {}
current_category = None

# Normalize category names: strip numeral prefix, keep consistent naming
def normalize_category(raw):
    # Remove leading "一、"/"十、" etc.
    name = re.sub(r"^(十|[一二三四五六七八九])、", "", raw)
    # Normalize spaces: "IPO 与资本运作" -> "IPO与资本运作", "互动与 ESG" -> "互动与ESG"
    name = name.replace(" ", "")
    return name

for line in md.splitlines():
    # Match ## 一、... or ## 十、...
    m = re.match(r"^##\s+(.+)", line)
    if m and ("、" in m.group(1)):
        current_category = normalize_category(m.group(1))
        continue
    # Match table rows with function names in backticks
    row_m = re.search(r"`([a-z_][a-z0-9_]+)`", line)
    if row_m and current_category:
        func_name = row_m.group(1)
        if func_name not in func_to_category:
            func_to_category[func_name] = current_category

print(f"Parsed {len(func_to_category)} function -> category mappings")

# Verify all 10 categories present
cats = set(func_to_category.values())
print(f"Categories found ({len(cats)}):")
for c in sorted(cats):
    count = sum(1 for v in func_to_category.values() if v == c)
    print(f"  {c}: {count}")

# ---- Step 2: Load interface data and test results ----
with open(INTERFACES_PATH, "r", encoding="utf-8") as f:
    interfaces = json.load(f)
with open(RESULTS_PATH, "r", encoding="utf-8") as f:
    results = json.load(f)

status_by_func = {r["func"]: r.get("status", "不可用") for r in results}

# ---- Step 3: Filter to available only, attach category ----
available = []
missing_cat = []
for iface in interfaces:
    if status_by_func.get(iface["func"]) != "可用":
        continue
    cat = func_to_category.get(iface["func"])
    if cat is None:
        missing_cat.append(iface["func"])
        cat = "未分类"
    params = iface.get("params_raw") or "null"
    if params == "" or params.lower() == "null":
        params = "null"
    available.append({
        "category": cat,
        "name": iface["name"],
        "data_source": iface["data_source"],
        "func": iface["func"],
        "url": iface["url"],
        "params": params,
    })

if missing_cat:
    print(f"\nWARNING: {len(missing_cat)} available functions not found in classification doc:")
    for fn in missing_cat:
        print(f"  {fn}")

# Sort by category order (as they appear in the doc), then by original order
cat_order = ["市场总览与统计", "标的列表与基础信息", "行情数据", "财务与估值",
             "股东与股本变动", "资金与筹码", "IPO与资本运作", "机构与研究",
             "公告与事件异动", "市场情绪、互动与ESG", "未分类"]
cat_index = {c: i for i, c in enumerate(cat_order)}
available.sort(key=lambda x: (cat_index.get(x["category"], 99), x["func"]))

# ---- Step 4: Generate Excel ----
wb = Workbook()
ws = wb.active
ws.title = "AkShare 股票接口测试"

headers = ["序号", "类别", "接口名称", "数据源", "接口", "目标地址", "输入参数", "可用性"]
ws.append(headers)

header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ok_fill = PatternFill("solid", fgColor="C6EFCE")
cell_font = Font(name="微软雅黑", size=10)
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Category color banding for readability
cat_colors = {
    "市场总览与统计": "E8F0FE",
    "标的列表与基础信息": "FCE8E8",
    "行情数据": "FEF7E0",
    "财务与估值": "E6F4EA",
    "股东与股本变动": "F3E8FD",
    "资金与筹码": "E0F7FA",
    "IPO与资本运作": "FFF3E0",
    "机构与研究": "F1F8E9",
    "公告与事件异动": "FCE4EC",
    "市场情绪、互动与ESG": "EDE7F6",
    "未分类": "F5F5F5",
}

for i, r in enumerate(available, 1):
    row = [i, r["category"], r["name"], r["data_source"], r["func"], r["url"], r["params"], "可用"]
    ws.append(row)
    row_idx = i + 1
    cat_fill = PatternFill("solid", fgColor=cat_colors.get(r["category"], "F5F5F5"))
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font
        cell.border = border
        if col_idx in (1, 8):
            cell.alignment = center_align
        else:
            cell.alignment = cell_align
    # Color the category cell with its category color
    ws.cell(row=row_idx, column=2).fill = cat_fill
    # Green for 可用
    ws.cell(row=row_idx, column=8).fill = ok_fill

widths = [6, 18, 34, 14, 36, 50, 26, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 24

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(available)+1}"

wb.save(OUT_PATH)

# Summary
print(f"\nSaved: {OUT_PATH}")
print(f"Total available: {len(available)}")
print(f"\n按类别统计:")
from collections import Counter
cat_counts = Counter(r["category"] for r in available)
for c in cat_order:
    if c in cat_counts:
        print(f"  {c}: {cat_counts[c]}")
