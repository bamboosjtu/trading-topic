"""Generate the final akshare.api.xlsx report from test results."""
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

RESULTS_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\.workbuddy\scratch\results.json"
OUT_PATH = r"C:\Users\theTruth\Documents\projects\vibe-working\trading-topic\akshare.api.xlsx"

with open(RESULTS_PATH, "r", encoding="utf-8") as f:
    results = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "AkShare 股票接口测试"

headers = ["序号", "接口名称", "数据源", "接口", "目标地址", "输入参数", "可用性"]
ws.append(headers)

# Header style
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, _ in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

ok_fill = PatternFill("solid", fgColor="C6EFCE")
fail_fill = PatternFill("solid", fgColor="FFC7CE")
cell_font = Font(name="微软雅黑", size=10)
cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

for i, r in enumerate(results, 1):
    params = r.get("params_raw", "null") or "null"
    if params == "" or params.lower() == "null":
        params = "null"
    row = [
        i,
        r.get("name", ""),
        r.get("data_source", ""),
        r.get("func", ""),
        r.get("url", ""),
        params,
        r.get("status", "不可用"),
    ]
    ws.append(row)
    row_idx = i + 1
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.font = cell_font
        cell.border = border
        if col_idx in (1, 7):  # 序号, 可用性
            cell.alignment = center_align
        else:
            cell.alignment = cell_align
    # color the status cell
    status_cell = ws.cell(row=row_idx, column=7)
    if status_cell.value == "可用":
        status_cell.fill = ok_fill
    else:
        status_cell.fill = fail_fill

# Column widths
widths = [6, 34, 16, 38, 52, 28, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze header row
ws.freeze_panes = "A2"
# Row height for header
ws.row_dimensions[1].height = 24

wb.save(OUT_PATH)

total = len(results)
ok = sum(1 for r in results if r.get("status") == "可用")
print(f"Saved: {OUT_PATH}")
print(f"Total: {total}, 可用: {ok}, 不可用: {total - ok}")
